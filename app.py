from selenium import webdriver
from selenium.webdriver.chrome.options import Options

from time import sleep
from selenium.webdriver.common.by import By

from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.webdriver.support.ui import WebDriverWait

from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains

import openpyxl
from openpyxl.styles import Font

class Car:
    def __init__(self):
        chrome_options = Options()
        chrome_options.add_argument("--ignore-certificate-errors")
        chrome_options.add_argument("--disable-logging")

        self.driver = webdriver.Chrome(options=chrome_options)
        self.link = "https://www.olx.com.br/autos-e-pecas/carros-vans-e-utilitarios/estado-pe/grande-recife/outras-cidades/caruaru?re=2025&rs=2023"

    def main(self):
        print('App Inicializado')
        try:
            self.initialize()
            sleep(2)
            self.scraping()
            sleep(2)
            self.create_sheet()
            sleep(2)
        finally:
            self.driver.quit()

    def initialize(self):
        self.driver.get(self.link)
        sleep(2)
        self.driver.maximize_window()
        sleep(2)
        self.driver.refresh()
        sleep(10)

    def safe_find_element(self, by, path, default="N/A"):
        try:
            element = self.driver.find_element(by, path)
            return element.text
        except NoSuchElementException:
            return default

    def scraping(self):
        global array_name, array_price, array_mileage, array_color, array_cylindrical, array_replacement

        array_name = []
        array_price = []
        array_mileage = []
        array_color = []
        array_cylindrical = []
        array_replacement = []

        counter = 1

        print('Inicializando o scraping')

        while True:
            xpaths = {
                'name': f'/html/body/div/div/main/div[2]/div/main/div[7]/section[{counter}]/div[2]/div[1]/div[1]/a',
                'price': f'/html/body/div/div/main/div[2]/div/main/div[7]/section[{counter}]/div[2]/div[1]/div[2]/h3',
                'mileage': f'/html/body/div/div/main/div[2]/div/main/div[7]/section[{counter}]/div[2]/div[1]/div[1]/ul[1]/li[1]/span',
                'color': f'/html/body/div/div/main/div[2]/div/main/div[7]/section[{counter}]/div[2]/div[1]/div[1]/ul[1]/li[2]/span',
                'cylindrical': f'/html/body/div/div/main/div[2]/div/main/div[7]/section[{counter}]/div[2]/div[1]/div[1]/ul[1]/li[3]/span',
                'replacement': f'/html/body/div/div/main/div[2]/div/main/div[7]/section[{counter}]/div[2]/div[1]/div[1]/ul[2]/li/div/span'
            }

            name = self.safe_find_element(By.XPATH, xpaths['name'])
            price = self.safe_find_element(By.XPATH, xpaths['price'])
            mileage = self.safe_find_element(By.XPATH, xpaths['mileage'])
            color = self.safe_find_element(By.XPATH, xpaths['color'])
            cylindrical = self.safe_find_element(By.XPATH, xpaths['cylindrical'])
            replacement = self.safe_find_element(By.XPATH, xpaths['replacement'])

            if name != "N/A" or price != "N/A" or mileage != "N/A" or color != "N/A" or cylindrical != "N/A" or replacement != "N/A":
                print(f'{name} | {price} | {mileage} | {color} | {cylindrical} | {replacement}')
                
                array_name.append(name)
                array_price.append(price)
                array_mileage.append(mileage)
                array_color.append(color)
                array_cylindrical.append(cylindrical)
                array_replacement.append(replacement)

            if name == "N/A" and price == "N/A" and counter > 1:
                try:
                    WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.LINK_TEXT, 'Próxima página')))
                    next_button = self.driver.find_element(By.LINK_TEXT, 'Próxima página')
                    
                    actions = ActionChains(self.driver)
                    actions.move_to_element(next_button).perform()
                    
                    next_button.click()
                    sleep(3)  
                except TimeoutException:
                    print('Scraping Finalizado.')
                    break
                except Exception as e:
                    print(f'Ocorreu um erro inesperado: {e}.')
                    break

            counter += 1

    def create_sheet(self):
        sheet = openpyxl.Workbook()
        car = sheet.active
        car.title = 'Carros'

        car['A1'] = 'CARRO'
        car['A1'].font = Font(bold=True)

        car['B1'] = 'PREÇO'
        car['B1'].font = Font(bold=True)

        car['C1'] = 'QUILOMETRAGEM'
        car['C1'].font = Font(bold=True)

        car['D1'] = 'COR'
        car['D1'].font = Font(bold=True)

        car['E1'] = 'CILINDRADAS'
        car['E1'].font = Font(bold=True)

        car['F1'] = 'ACEITA TROCA?'
        car['F1'].font = Font(bold=True)

        for index, (name, price, mileage, color, cylindrical, replacement) in enumerate(zip(array_name, array_price, array_mileage, array_color, array_cylindrical, array_replacement), start=2):
            car.cell(column=1, row=index, value=name)
            car.cell(column=2, row=index, value=price)
            car.cell(column=3, row=index, value=mileage)
            car.cell(column=4, row=index, value=color)
            car.cell(column=5, row=index, value=cylindrical)
            car.cell(column=6, row=index, value=replacement)

    
        for col in car.columns:
            max_length = 0
            column = col[0].column_letter  
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            car.column_dimensions[column].width = adjusted_width

        sheet.save('sheet_price_cars.xlsx')
        print('Planilha salva com sucesso!')

car = Car()
car.main()
